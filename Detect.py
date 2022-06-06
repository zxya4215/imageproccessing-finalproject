import cv2 as cv
import openpyxl
import time

from Settings import *
from Hash import GetImageHash

wb = openpyxl.load_workbook(HASH_EXCEL_PATH)
sheet = wb.active


def HammingDistance(tempImgHash: str, imgHash: str) -> int:
    w = 0
    for i in range(len(tempImgHash)):
        w += bin(int(tempImgHash[i], HASH_BASE) ^
                 int(imgHash[i], HASH_BASE)).count("1")
    return w


def ExcRead(row: int):
    cardName = sheet.cell(row, 1).value
    dct = sheet.cell(row, 2).value
    return cardName, dct

