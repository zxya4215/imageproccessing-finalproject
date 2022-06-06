from openpyxl import Workbook
from os import listdir
from os.path import isfile, join

from Settings import IMAGE_DB_PATH, HASH_EXCEL_PATH
from Hash import GetImageHash

def export():
    wb = Workbook()
    sheet = wb.create_sheet("卡片資料庫", 0)
    sheet['A1'] = "卡片編號"
    sheet['B1'] = "DCT值"

    i = 2
    for f in listdir(IMAGE_DB_PATH):
        fullpath = join(IMAGE_DB_PATH, f)
        if isfile(fullpath):
            sheet.cell(row=i, column=1).value = f
            sheet.cell(row=i, column=2).value = GetImageHash(fullpath)
            i += 1
    wb.save(HASH_EXCEL_PATH)

export()